import openpyxl
import os


class my_pyxl():
    def __init__(self, filename=None):
        self.filename = filename

    def add_write(self, table, filename='', sheet_name='sheet1'):
        """写文件xlsx"""
        filename = filename or self.filename
        if not filename:
            print('没有指定文件')
            return "没有指定文件"
        if os.path.exists(filename):
            print("已存在同名文件是否覆盖写入,不覆盖则为追加写入")
            is_rewrite = self.yes_input()
        if not is_rewrite:
            table += self.read_excel_xlsx(filename, sheet_name)
        behave_str = "\033[1;33m覆盖\033[0m" if is_rewrite else "\033[1;36m写入\033[0m"
        try:
            self.write_excel_xlsx(filename, sheet_name, table)
            print("\033[1;33m{}\033[0m 表格{}数据\033[1;32m成功\033[0m！".format(
                filename, behave_str))
            return "{} 表格{}数据成功！".format(filename, behave_str)
        except Exception as e:
            print("\033[1;33m{}\033[0m 表格{}数据\033[1;31m失败\033[0m！".format(
                filename, behave_str))
            print(type(e),e)
            return "{} 表格{}数据失败！".format(filename, behave_str)

    @staticmethod
    def write_excel_xlsx(filename, sheet_name, table):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(len(table)):
            for j in range(len(table[i])):
                sheet.cell(row=i+1, column=j+1, value=str(table[i][j]))
        workbook.save(filename)

    @staticmethod
    def read_excel_xlsx(filename, sheet_name='sheet1'):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]
        table = []
        for row in sheet.rows:
            table.append([cell.value for cell in row])
        return table

    def yes_input(self):
        while True:
            input_str = input('yes or no >')
            input_str_lower = input_str.lower()
            if 'yes' == input_str_lower or 'y' == input_str_lower:
                return True
            elif 'no' == input_str_lower or 'n' == input_str_lower:
                return False
            else:
                print('input input yes or no not any other key')


if __name__ == '__main__':
    filename = 'demo.xlsx'
    table = [
        [1, 2, 3],
        [2, 3, 4]
    ]
    my_pyxl(filename).add_write(table)
